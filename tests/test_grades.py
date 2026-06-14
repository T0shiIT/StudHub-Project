import io
import pytest
import requests
from conftest import BASE_URL, register_and_login, make_user_payload
 
#создане оценок
def _get_csrf(session: requests.Session) -> str | None:
    """Получаем CSRF-токен из куки Spring после любого GET-запроса."""
    session.get(f"{BASE_URL}/api/user", timeout=10)
    return session.cookies.get("XSRF-TOKEN")
 
#просмотр оценок
class TestGetGrades:
 
    def test_student_gets_own_grades_200(self, student):
        r = student["session"].get(f"{BASE_URL}/api/grades", timeout=10)
        assert r.status_code == 200
        assert isinstance(r.json(), list)
 
    def test_unauthenticated_get_grades_401(self):
        r = requests.get(f"{BASE_URL}/api/grades", timeout=10)
        assert r.status_code == 401
 
    def test_teacher_requires_group_param(self, teacher):
        """Teacher без group → 400."""
        r = teacher["session"].get(f"{BASE_URL}/api/grades", timeout=10)
        assert r.status_code == 400
 
    def test_teacher_with_group_param_200(self, teacher):
        r = teacher["session"].get(
            f"{BASE_URL}/api/grades",
            params={"group": "ПИ-2025"},
            timeout=10,
        )
        assert r.status_code == 200
        assert isinstance(r.json(), list)
 
    def test_teacher_filter_by_subject(self, teacher):
        r = teacher["session"].get(
            f"{BASE_URL}/api/grades",
            params={"group": "ПИ-2025", "subject": "Математика"},
            timeout=10,
        )
        assert r.status_code == 200
        body = r.json()
        # Если оценок нет — пустой список, не ошибка
        assert isinstance(body, list)
        for g in body:
            assert g.get("subject") == "Математика"
 
    def test_admin_with_group_param_200(self, admin):
        r = admin["session"].get(
            f"{BASE_URL}/api/grades",
            params={"group": "ПИ-2025"},
            timeout=10,
        )
        assert r.status_code == 200
 
#обновление оценок
class TestUpdateGrade:
 
    def _create_grade_via_upload(self, uploader_session: requests.Session, student_email: str) -> int | None:
        """
        Генерирует минимальный CSV-подобный xlsx в памяти и загружает через /api/grades/upload.
        Возвращает id первой созданной оценки или None если upload не поддерживается/провалился.
        
        Реальный xlsx создаётся через openpyxl; если библиотека не установлена — пропускаем.
        """
        try:
            import openpyxl
        except ImportError:
            return None
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["student_email", "subject", "grade", "date"])
        ws.append([student_email, "Физика", "5", "2025-01-15"])
 
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
 
        csrf = _get_csrf(uploader_session)
        files = {"file": ("grades.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        headers = {"X-XSRF-TOKEN": csrf} if csrf else {}
        r = uploader_session.post(
            f"{BASE_URL}/api/grades/upload",
            files=files,
            headers=headers,
            timeout=30,
        )
        if r.status_code != 200:
            return None
 
        # Получаем id через GET оценок студента (если есть сессия студента)
        return r.json().get("processed", 0)
 
    def test_patch_nonexistent_grade_404(self, teacher):
        sess = teacher["session"]
        csrf = _get_csrf(sess)
        r = sess.patch(
            f"{BASE_URL}/api/grades/999999999",
            json={"grade": "5"},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=10,
        )
        assert r.status_code == 404
 
    def test_student_cannot_patch_grade(self, student):
        sess = student["session"]
        csrf = _get_csrf(sess)
        r = sess.patch(
            f"{BASE_URL}/api/grades/1",
            json={"grade": "5"},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=10,
        )
        # студент получает 403 или 404
        assert r.status_code in (403, 404)
 
    def test_unauthenticated_cannot_patch(self):
        r = requests.patch(f"{BASE_URL}/api/grades/1", json={"grade": "5"}, timeout=10)
        assert r.status_code == 401
 
    @pytest.mark.integration
    def test_teacher_can_upload_and_update_grade(self, teacher, student):
        """
        Интеграционный тест: загружаем оценку, потом обновляем через PATCH.
        Требует openpyxl: pip install openpyxl
        """
        pytest.importorskip("openpyxl")
        import openpyxl
 
        #upload
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["student_email", "subject", "grade", "date"])
        ws.append([student["email"], "Химия", "3", "2025-03-10"])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
 
        csrf = _get_csrf(teacher["session"])
        upload_r = teacher["session"].post(
            f"{BASE_URL}/api/grades/upload",
            files={"file": ("g.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=30,
        )
        assert upload_r.status_code == 200, f"Upload failed: {upload_r.text}"
        assert upload_r.json()["processed"] >= 1
 
        #найти grade_id через GET студента
        grades_r = student["session"].get(f"{BASE_URL}/api/grades", timeout=10)
        assert grades_r.status_code == 200
        grades = grades_r.json()
        target = next((g for g in grades if g["subject"] == "Химия"), None)
        assert target is not None, "Оценка по Химии не найдена после upload"
 
        grade_id = target["id"]
 
        #PATCH
        csrf = _get_csrf(teacher["session"])
        patch_r = teacher["session"].patch(
            f"{BASE_URL}/api/grades/{grade_id}",
            json={"grade": "5"},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=10,
        )
        assert patch_r.status_code == 200
        assert patch_r.json()["grade"] == "5"
 
        #проверяем что студент видит обновлённую оценку
        grades_after = student["session"].get(f"{BASE_URL}/api/grades", timeout=10).json()
        updated = next((g for g in grades_after if g["id"] == grade_id), None)
        assert updated["grade"] == "5"
 
#отправка оценок
class TestUploadGrades:
 
    def test_student_cannot_upload_grades(self, student):
        pytest.importorskip("openpyxl")
        import openpyxl
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["student_email", "subject", "grade", "date"])
        ws.append([student["email"], "Физика", "4", "2025-02-01"])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
 
        csrf = _get_csrf(student["session"])
        r = student["session"].post(
            f"{BASE_URL}/api/grades/upload",
            files={"file": ("g.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=20,
        )
        assert r.status_code == 403
 
    def test_upload_empty_file_returns_400(self, teacher):
        csrf = _get_csrf(teacher["session"])
        r = teacher["session"].post(
            f"{BASE_URL}/api/grades/upload",
            files={"file": ("empty.xlsx", io.BytesIO(b""), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"X-XSRF-TOKEN": csrf} if csrf else {},
            timeout=10,
        )
        assert r.status_code == 400
 
    def test_upload_without_auth_returns_401(self):
        pytest.importorskip("openpyxl")
        import openpyxl
 
        wb = openpyxl.Workbook()
        wb.active.append(["student_email", "subject", "grade", "date"])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
 
        r = requests.post(
            f"{BASE_URL}/api/grades/upload",
            files={"file": ("g.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=10,
        )
        assert r.status_code == 401
 